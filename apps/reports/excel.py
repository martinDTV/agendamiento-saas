"""
Excel export of the appointment report, with native Excel charts.

Builds a multi-sheet workbook from the report payload produced by
services.build_report(). Charts are real Excel chart objects (editable in
Excel), not images.
"""
from io import BytesIO

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .services import STATUS_LABELS, WEEKDAY_LABELS

# Brand colors (no leading #, openpyxl wants ARGB/RGB hex).
NAVY = '274C77'
VIOLET = 'A78BFA'
HEADER_FILL = PatternFill('solid', fgColor=NAVY)
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
TITLE_FONT = Font(color=NAVY, bold=True, size=16)
LABEL_FONT = Font(color='475569', size=10)


def _style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')


def _autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_workbook(report, tenant_name):
    wb = Workbook()

    # ── Sheet 1: Resumen ──────────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Resumen'
    ws['A1'] = f'Reporte de citas — {tenant_name}'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = f"Periodo: {report['from_date']} a {report['to_date']}"
    ws['A2'].font = LABEL_FONT

    ws['A4'] = 'Métrica'
    ws['B4'] = 'Valor'
    _style_header(ws, 4, 2)
    summary = [
        ('Total de citas', report['total']),
        ('Ingreso estimado (MXN)', report['revenue_total']),
        ('Tasa de cancelación (%)', report['cancellation_rate']),
        ('Tasa de finalización (%)', report['completion_rate']),
        ('Pendientes', report['by_status']['pending']),
        ('Confirmadas', report['by_status']['confirmed']),
        ('Canceladas', report['by_status']['cancelled']),
        ('Completadas', report['by_status']['completed']),
    ]
    for i, (label, value) in enumerate(summary, start=5):
        ws.cell(row=i, column=1, value=label)
        ws.cell(row=i, column=2, value=value)
    _autosize(ws, [28, 18])

    # Pie chart of statuses
    status_start = 5 + len(summary) + 1
    ws.cell(row=status_start, column=1, value='Estado')
    ws.cell(row=status_start, column=2, value='Citas')
    _style_header(ws, status_start, 2)
    status_rows = [
        ('Pendiente', report['by_status']['pending']),
        ('Confirmada', report['by_status']['confirmed']),
        ('Cancelada', report['by_status']['cancelled']),
        ('Completada', report['by_status']['completed']),
    ]
    for j, (label, value) in enumerate(status_rows, start=status_start + 1):
        ws.cell(row=j, column=1, value=label)
        ws.cell(row=j, column=2, value=value)
    pie = PieChart()
    pie.title = 'Citas por estado'
    labels = Reference(ws, min_col=1, min_row=status_start + 1, max_row=status_start + len(status_rows))
    data = Reference(ws, min_col=2, min_row=status_start, max_row=status_start + len(status_rows))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, 'D4')

    # ── Sheet 2: Por doctor ───────────────────────────────────────────────
    _table_with_bar(
        wb.create_sheet('Por doctor'),
        report['by_doctor'],
        headers=['Doctor', 'Citas', 'Ingreso (MXN)'],
        keys=['name', 'total', 'revenue'],
        chart_title='Citas por doctor',
        widths=[32, 12, 16],
    )

    # ── Sheet 3: Por servicio ─────────────────────────────────────────────
    _table_with_bar(
        wb.create_sheet('Por servicio'),
        report['by_service'],
        headers=['Servicio', 'Citas', 'Ingreso (MXN)'],
        keys=['name', 'total', 'revenue'],
        chart_title='Citas por servicio',
        widths=[32, 12, 16],
    )

    # ── Sheet 4: Tendencia ────────────────────────────────────────────────
    ws_t = wb.create_sheet('Tendencia')
    ws_t['A1'] = 'Citas en el tiempo'
    ws_t['A1'].font = TITLE_FONT
    ws_t['A3'] = 'Fecha'
    ws_t['B3'] = 'Citas'
    _style_header(ws_t, 3, 2)
    for i, point in enumerate(report['trend'], start=4):
        ws_t.cell(row=i, column=1, value=point['date'])
        ws_t.cell(row=i, column=2, value=point['total'])
    _autosize(ws_t, [16, 12])
    if report['trend']:
        line = LineChart()
        line.title = 'Tendencia de citas'
        n = len(report['trend'])
        data = Reference(ws_t, min_col=2, min_row=3, max_row=3 + n)
        cats = Reference(ws_t, min_col=1, min_row=4, max_row=3 + n)
        line.add_data(data, titles_from_data=True)
        line.set_categories(cats)
        line.height = 8
        line.width = 18
        ws_t.add_chart(line, 'D3')

    # ── Sheet 5: Ocupación (heatmap data) ─────────────────────────────────
    ws_o = wb.create_sheet('Ocupación')
    ws_o['A1'] = 'Ocupación por día y hora'
    ws_o['A1'].font = TITLE_FONT
    ws_o['A3'] = 'Día'
    ws_o['B3'] = 'Hora'
    ws_o['C3'] = 'Citas'
    _style_header(ws_o, 3, 3)
    for i, cell in enumerate(report['occupancy'], start=4):
        ws_o.cell(row=i, column=1, value=WEEKDAY_LABELS[cell['weekday']])
        ws_o.cell(row=i, column=2, value=f"{cell['hour']:02d}:00")
        ws_o.cell(row=i, column=3, value=cell['total'])
    _autosize(ws_o, [14, 10, 10])

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _table_with_bar(ws, rows, headers, keys, chart_title, widths):
    ws['A1'] = chart_title
    ws['A1'].font = TITLE_FONT
    header_row = 3
    for c, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=c, value=h)
    _style_header(ws, header_row, len(headers))
    for i, row in enumerate(rows, start=header_row + 1):
        for c, key in enumerate(keys, start=1):
            ws.cell(row=i, column=c, value=row.get(key))
    _autosize(ws, widths)

    if rows:
        bar = BarChart()
        bar.type = 'col'
        bar.title = chart_title
        n = len(rows)
        data = Reference(ws, min_col=2, min_row=header_row, max_row=header_row + n)
        cats = Reference(ws, min_col=1, min_row=header_row + 1, max_row=header_row + n)
        bar.add_data(data, titles_from_data=True)
        bar.set_categories(cats)
        bar.height = 9
        bar.width = 18
        ws.add_chart(bar, f'{get_column_letter(len(headers) + 2)}3')
